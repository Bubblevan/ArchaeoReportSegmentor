import openai
from openai import OpenAI
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
# from zhipuai import ZhipuAI

client = OpenAI(api_key = "sk-ba5e139df4734a05a95519a5e70995fa", base_url="https://api.deepseek.com/")


def split_text_into_chunks(text, max_tokens=2048):
    chunks = []
    start = 0
    while start < len(text):
        end = min(start + max_tokens, len(text))
        chunks.append(text[start:end])
        start = end
    return chunks

if __name__ == '__main__':
    # 从文本文件中读取内容
    with open('./ocr_result/ocr_output.txt', 'r', encoding='utf-8') as file:
        input_text = file.read()

    # 将文本按一定量的tokens分段
    max_tokens_per_chunk = 500
    chunks = split_text_into_chunks(input_text, max_tokens=max_tokens_per_chunk)

    if os.path.exists('./llm_result/info.xlsx'):
        wb = openpyxl.load_workbook('./llm_result/info.xlsx')
        ws = wb.active
        # 获取当前工作表的最后一行，以便追加数据
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        start_row = 1   # 如果是新文件，从第一行开始写入

    # 设置列宽和表头
    if start_row == 1:
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20

        # 写入表头
        header = ["遗迹号", "序号", "尺寸", "外形描述", "颜色", "器型1", "器型2", "材质", "完整程度", "图注"]
        for col_num, col_name in enumerate(header, 1):
            header_cell = ws.cell(row=1, column=col_num, value=col_name)
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1      # 数据从表头下一行开始写入

    # 处理每一个文本块
    for chunk in chunks:
        prompt_text = f"{chunk}\n\n\
        作为一个专业的考古学者，请帮我整理一份以上所有器物的考古器物卡片的表格。\n\
        这份表格需要按照遗迹号、序号、尺寸、外形描述、颜色、器型1、器型2、材质、完整程度、图注的顺序来整理。在整理时，请遵循以下提取要求：\n\
        遵循以下提取要求：\n\
        遗迹号：每个器物号冒号前的数字及字母\n\
        序号：每个器物号冒号后的数字\n\
        尺寸：直接从原句中提取尺寸信息，无需修改。如果尺寸中有多个数值，请使用顿号（、）进行连接。\n\
        外形描述：直接从原句中提取，无需进行任何总结或概括。\n\
        器型1：指该器物是属于玉器/陶器/石器/漆器其中某一大类\n\
        器型2：指的是器物号之前标注的名称如：罐，环，壶，鼎等。如果某个器物没有特定的标注名称，则与上一个器物的器型2相同。\n\
        材质：器物的材质，但请注意，器型与材质是不同的概念。注意区分颜色和材质\n\
        完整程度：即使是残碎的器物，也请务必记录在表格中。\n\
        图注为器物信息末尾的括号里的内容，直接提取无需修改。请基于我所提供的考古器物信息来进行整理。提供一个格式清晰、易于理解的考古器物卡片表格，并将生成的所有表格都展示出来。表格中应包含上述提到的所有要素，并按顺序排列。"
        
        # 调用DeepSeek API进行文本生成
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "user", "content": prompt_text}
            ],
            temperature=0.95,
            top_p=0.7
        )
        # 获取生成的文本结果
        generated_text = response.choices[0].message.content.strip()
        
        # 分割器物段落
        artifact_paragraphs = [para.strip() for para in generated_text.split("\n") if para.strip()]

        # 将器物段落添加到Excel表格
        for para in artifact_paragraphs:
            if "遗迹号" in para:
                # 当遇到新的遗迹号时，重置start_row为当前行号+1
                start_row = ws.max_row + 1
                continue
            if para:
                artifact_info = [info.strip() for info in para.split("|") if info.strip()]
                for col_num, info in enumerate(artifact_info, 1):
                    ws.cell(row=start_row, column=col_num, value=info)
                start_row += 1

    wb.save('./llm_result/info.xlsx')
    print("Artifact information saved to 'info.xlsx'")
